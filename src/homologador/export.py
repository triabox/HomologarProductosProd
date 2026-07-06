"""Exportación de resultados: Excel (datos completos para procesar) y PDF (formato web).

- Excel: workbook multi-hoja con TODOS los datos de la corrida (resumen vs objetivos,
  categorías, detalle por producto, resultados por campo en formato largo, atributos
  faltantes, conteos por categoría e histórico de corridas).
- PDF: render del HTML del panel principal con WeasyPrint (mismo formato que la web).
  Si WeasyPrint no está disponible (faltan libs de sistema), se omite con aviso.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Config
from .engine import Engine
from .report import _duration
from .stats import global_summary
from .storage import Storage

_HDR_FILL = PatternFill("solid", fgColor="1E293B")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_BAD_FONT = Font(color="CC0000", bold=True)
_OK_FONT = Font(color="1A7F37")


def _sheet(wb: Workbook, title: str, headers: list[str], widths: Optional[list[int]] = None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(vertical="center")
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.freeze_panes = "A2"
    return ws


def export_excel(cfg: Config, storage: Storage, run_id: int, out_path: Path) -> Path:
    """Genera el .xlsx con toda la información de la corrida, lista para procesar."""
    labels = Engine(cfg).comparator_labels()
    goals = cfg.get("goals", {}) or {}
    summary = global_summary(storage, run_id)
    run = storage.get_run(run_id)
    _, dur = _duration(run["started_at"], run["finished_at"])

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Resumen ----
    ws = _sheet(wb, "Resumen", ["Indicador", "Valor", "Meta", "Estado"], [30, 16, 10, 12])
    ws.append(["Corrida", f"#{run_id}", "", ""])
    ws.append(["Inicio", run["started_at"], "", ""])
    ws.append(["Duración", dur, "", ""])
    ws.append(["Productos comparados", summary.products_compared, "", ""])
    ws.append(["Con par en VTEX", summary.vtex_found, "", ""])
    ws.append(["Cobertura SKU (%)", summary.coverage_pct, "", ""])
    ws.append(["Score homologación", summary.avg_score, "", ""])
    for key, label in labels.items():
        val = summary.field_ok.get(key, 0.0)
        goal = goals.get(key)
        estado = "" if goal is None else ("CUMPLE" if val >= goal else "DEBAJO")
        row = ws.append([f"% OK {label}", val, goal or "", estado])
        c = ws.cell(row=ws.max_row, column=4)
        c.font = _OK_FONT if estado == "CUMPLE" else (_BAD_FONT if estado else c.font)

    # ---- Categorías (validación) ----
    ws = _sheet(wb, "Categorías",
                ["Categoría", "Muestra", "En VTEX", "Score",
                 *[f"% OK {v}" for v in labels.values()], "URL CoRD", "URL VTEX"],
                [34, 9, 9, 9] + [12] * len(labels) + [45, 45])
    for c in storage.category_stats(run_id):
        fo = json.loads(c["field_ok_json"] or "{}")
        ws.append([c["category_name"], c["sampled"], c["vtex_found"], c["avg_score"],
                   *[fo.get(k, "") for k in labels], c["cord_url"], c["vtex_url"]])

    # ---- Detalle por producto (ancho: un flag OK por campo) ----
    frs = storage.field_results(run_id)
    by_sku: dict[str, dict] = {}
    for fr in frs:
        by_sku.setdefault(fr["sku"], {})[fr["field"]] = fr
    ws = _sheet(wb, "Productos",
                ["SKU", "Categoría", "Score", "En VTEX",
                 *[f"OK {v}" for v in labels.values()], "URL CoRD", "URL VTEX", "Error"],
                [12, 30, 9, 9] + [12] * len(labels) + [45, 45, 25])
    for p in storage.product_results(run_id):
        flds = by_sku.get(p["sku"], {})
        ws.append([p["sku"], p["category_name"], p["score"], "sí" if p["vtex_found"] else "NO",
                   *[("sí" if flds[k]["ok"] else "NO") if k in flds else "" for k in labels],
                   p["cord_url"], p["vtex_url"], p["error"] or ""])

    # ---- Resultados por campo (largo: toda la data cruda) ----
    ws = _sheet(wb, "Campos",
                ["SKU", "Campo", "OK", "Score", "Severidad", "Valor CoRD", "Valor VTEX", "Detalle"],
                [12, 18, 6, 8, 14, 30, 30, 60])
    for fr in frs:
        ws.append([fr["sku"], labels.get(fr["field"], fr["field"]), "sí" if fr["ok"] else "NO",
                   fr["score"], fr["severity"], fr["cord_value"], fr["vtex_value"], fr["detail"]])
        if not fr["ok"]:
            ws.cell(row=ws.max_row, column=3).font = _BAD_FONT

    # ---- Atributos faltantes ----
    ws = _sheet(wb, "Atributos faltantes",
                ["Atributo", "Productos afectados", "Categorías", "SKUs"],
                [30, 18, 12, 80])
    for r in storage.attribute_gaps(run_id, "missing"):
        ws.append([r["label"], r["products"], r["categories"], r["skus"]])

    # ---- Conteos por categoría (acumulado más reciente) ----
    ws = _sheet(wb, "Conteos",
                ["Categoría", "Productos CoRD", "Productos VTEX", "Diferencia",
                 "Estado", "URL CoRD", "URL VTEX"],
                [34, 15, 15, 12, 14, 45, 45])
    for r in storage.latest_category_counts():
        cc, vc = r["cord_count"] or 0, r["vtex_count"] or 0
        estado = "igual" if cc == vc else ("CoRD>VTEX" if cc > vc else "faltan en CoRD")
        ws.append([r["category_name"], cc, vc, cc - vc, estado, r["cord_url"], r["vtex_url"]])
        if estado != "igual":
            ws.cell(row=ws.max_row, column=5).font = _BAD_FONT

    # ---- Histórico de corridas ----
    ws = _sheet(wb, "Histórico",
                ["Corrida", "Inicio", "Duración", "Productos", "Score", "Cobertura %",
                 *[f"% OK {v}" for v in labels.values()]],
                [9, 22, 10, 11, 9, 12] + [12] * len(labels))
    for r in storage.all_runs():
        s = global_summary(storage, r["id"])
        _, d = _duration(r["started_at"], r["finished_at"])
        ws.append([r["id"], r["started_at"], d, s.products_compared, s.avg_score,
                   s.coverage_pct, *[s.field_ok.get(k, "") for k in labels]])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# CSS extra para paginar bien el dashboard al imprimir
_PDF_CSS = """
@page { size: A4 landscape; margin: 10mm; background: #0f172a; }
body { background: #0f172a !important; }
table { font-size: 9px !important; }
tr, .goal, .sec { break-inside: avoid; }
details { display: block; }
details > * { display: block; }
"""


def export_pdf(html_path: Path, out_path: Path) -> Optional[Path]:
    """Convierte el HTML del dashboard a PDF (mismo formato que la web).

    Devuelve None si WeasyPrint no está disponible (dependencias de sistema).
    """
    try:
        from weasyprint import CSS, HTML  # import diferido: es opcional
    except Exception as e:
        print(f"[export] PDF omitido: WeasyPrint no disponible ({type(e).__name__}). "
              f"Instalá con: pip install 'homologador[pdf]' (+ libpango en el sistema)")
        return None
    out_path.parent.mkdir(parents=True, exist_ok=True)
    HTML(filename=str(html_path)).write_pdf(
        str(out_path), stylesheets=[CSS(string=_PDF_CSS)]
    )
    return out_path
